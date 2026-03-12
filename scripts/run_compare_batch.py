from __future__ import annotations

import json
import os
import random
import statistics
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

# Force UTF-8 console output on Windows terminals.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import matplotlib.pyplot as plt

from hq_dhfsp.algorithms.moead import MOEADConfig
from hq_dhfsp.algorithms.nsga2 import NSGA2Config
from hq_dhfsp.config import GlobalConfig
from hq_dhfsp.decoder import decode
from hq_dhfsp.instance import ProblemInstance
from hq_dhfsp.metrics.contribution import nondominated
from hq_dhfsp.metrics.hv import hv_approx
from hq_dhfsp.metrics.igd import igd
from hq_dhfsp.metrics.spacing import spacing
from hq_dhfsp.runner import RunnerConfig, run, run_single
from hq_dhfsp.viz.compare import plot_hv_compare
from hq_dhfsp.viz.gantt import plot_gantt
from hq_dhfsp.viz.pareto import plot_pareto
from hq_dhfsp.viz.pareto_compare import plot_pareto_compare
from hq_dhfsp.viz.progress import plot_progress

from openpyxl import Workbook

try:
    from scipy.stats import friedmanchisquare, wilcoxon
except Exception:  # pragma: no cover - optional dependency
    friedmanchisquare = None
    wilcoxon = None


# Stability switches for long batch runs on Windows.
ENABLE_RUN_PLOTS = False
ENABLE_CASE_SUMMARY_PLOTS = True
CASE_LIMIT = 5  # run first five scales in full batch mode

# 快速小规模测试：仅 30J3S2F 一例
QUICK_MEDIUM_TEST = False
MEDIUM_SPEC = {"J": 30, "F": 2, "S": 3}
MEDIUM_RUNS = 3
MEDIUM_MAX_ITERS = 50  # 小规模折中测试迭代数


def _setup_plot_style() -> None:
    plt.rcParams["font.family"] = "Times New Roman"
    plt.rcParams["font.size"] = 9
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["axes.linewidth"] = 1.5
    plt.rcParams["lines.linewidth"] = 1.6


def _plot_bar(
    title: str,
    labels: List[str],
    values: List[float],
    out_path: Path,
    ylabel: str,
    yerr: List[float] | None = None,
) -> None:
    _setup_plot_style()
    fig, ax = plt.subplots(figsize=(7.2, 4.2))
    x = list(range(len(labels)))
    ax.bar(x, values, yerr=yerr, capsize=3, color="#4C78A8")
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()
    fig.savefig(out_path, dpi=600)
    plt.close(fig)


def plot_hq_mechanism_case(metrics: dict, case_dir: Path, case_name: str) -> None:
    hq = metrics.get("HQ", {})
    if not hq:
        return
    cr = hq.get("cr", {}) or {}
    mode_counts = hq.get("mode_counts", {}) or {}
    op_counts = hq.get("op_counts", {}) or {}
    op_success = hq.get("op_success", {}) or {}

    algo_order = [("nsga", "NSGA-II"), ("moead", "MOEA/D"), ("hho", "HHO"), ("ig", "IG")]
    cr_vals = [float(cr.get(k, 0.0)) for k, _ in algo_order]
    cr_labels = [label for _, label in algo_order]
    _plot_bar(
        title=f"{case_name} - HQ贡献率(CR)",
        labels=cr_labels,
        values=cr_vals,
        out_path=case_dir / "hq_mech_cr.pdf",
        ylabel="CR",
    )

    mode_total = sum(mode_counts.values()) or 1
    mode_order = [("independent", "Independent"), ("cooperation", "Cooperation"), ("competition", "Competition")]
    mode_vals = [mode_counts.get(k, 0) / mode_total for k, _ in mode_order]
    mode_labels = [label for _, label in mode_order]
    _plot_bar(
        title=f"{case_name} - HQ模式占比",
        labels=mode_labels,
        values=mode_vals,
        out_path=case_dir / "hq_mech_mode.pdf",
        ylabel="Ratio",
    )

    op_order = [
        ("c1_elite_migration", "C1"),
        ("c2_rhythm_coop", "C2"),
        ("r1_struct_suppress", "R1"),
        ("r2_territorial_invade", "R2"),
    ]
    op_vals = []
    op_labels = []
    for k, label in op_order:
        cnt = op_counts.get(k, 0)
        if cnt > 0:
            op_vals.append(op_success.get(k, 0) / cnt)
            op_labels.append(label)
    if op_vals:
        _plot_bar(
            title=f"{case_name} - HQ算子成功率",
            labels=op_labels,
            values=op_vals,
            out_path=case_dir / "hq_mech_op_success.pdf",
            ylabel="Success Rate",
        )


def _quantile(values: List[float], q: float) -> float:
    if not values:
        return 0.0
    vals = sorted(values)
    idx = int(round(q * (len(vals) - 1)))
    return vals[max(0, min(idx, len(vals) - 1))]


def _classify_regions(objs: List[Tuple[float, float, float]]) -> List[str]:
    if not objs:
        return []
    mins = [min(o[i] for o in objs) for i in range(3)]
    maxs = [max(o[i] for o in objs) for i in range(3)]
    norm = []
    for o in objs:
        n = []
        for i in range(3):
            rng = maxs[i] - mins[i]
            n.append((o[i] - mins[i]) / (rng + 1e-9))
        norm.append(n)
    thresholds = [_quantile([n[i] for n in norm], 0.1) for i in range(3)]
    sums = [sum(n) for n in norm]
    knee_thr = _quantile(sums, 0.2)
    regions = []
    for n, s in zip(norm, sums):
        if any(n[i] <= thresholds[i] for i in range(3)):
            regions.append("extreme")
        elif s <= knee_thr:
            regions.append("knee")
        else:
            regions.append("middle")
    return regions


def _collect_ref_front(algo_objs: Dict[str, List[Tuple[float, float, float]]]) -> List[Tuple[float, float, float]]:
    all_objs = []
    for objs in algo_objs.values():
        all_objs.extend(objs)
    return nondominated(all_objs)


def _contribution_by_region(
    algo_objs: Dict[str, List[Tuple[float, float, float]]]
) -> Dict[str, Dict[str, float]]:
    ref = _collect_ref_front(algo_objs)
    ref_regions = _classify_regions(ref)
    ref_map: Dict[Tuple[float, float, float], List[str]] = {}
    for o, r in zip(ref, ref_regions):
        ref_map.setdefault(o, []).append(r)
    total = len(ref) or 1
    contrib: Dict[str, Dict[str, float]] = {}
    for algo, objs in algo_objs.items():
        counts = {"extreme": 0, "middle": 0, "knee": 0}
        for o in objs:
            if o in ref_map:
                # if duplicates, count once per occurrence
                for r in ref_map[o]:
                    counts[r] += 1
        contrib[algo] = {k: v / total for k, v in counts.items()}
    return contrib


def _plot_cr_grouped_stacked(
    instance_names: List[str],
    cr_runs: Dict[str, List[Dict[str, Dict[str, float]]]],
    out_path: Path,
    sig_marks: Dict[str, Dict[str, bool]],
) -> None:
    _setup_plot_style()
    algos = ["HQ", "NSGA-II", "MOEA/D", "HHO"]
    regions = [("extreme", "///"), ("middle", "\\\\\\"), ("knee", "xxx")]
    n_inst = len(instance_names)
    width = 0.18
    x = list(range(n_inst))

    fig, ax = plt.subplots(figsize=(12, 4.2))
    for ai, algo in enumerate(algos):
        means = []
        stds = []
        stacks = {r[0]: [] for r in regions}
        for inst in instance_names:
            runs = cr_runs.get(inst, [])
            vals = [r.get(algo, {"extreme": 0, "middle": 0, "knee": 0}) for r in runs]
            totals = [v["extreme"] + v["middle"] + v["knee"] for v in vals]
            m, s = _mean_std(totals)
            means.append(m)
            stds.append(s)
            for r, _ in regions:
                stacks[r].append(_mean_std([v[r] for v in vals])[0])

        bottom = [0.0] * n_inst
        for r, hatch in regions:
            ax.bar(
                [xi + (ai - 1.5) * width for xi in x],
                stacks[r],
                width=width,
                bottom=bottom,
                color="white",
                edgecolor="black",
                hatch=hatch,
                label=r if ai == 0 else None,
            )
            bottom = [b + v for b, v in zip(bottom, stacks[r])]
        ax.errorbar(
            [xi + (ai - 1.5) * width for xi in x],
            means,
            yerr=stds,
            fmt="none",
            ecolor="black",
            elinewidth=1.2,
            capsize=2,
        )
        if algo == "HQ":
            for idx, inst in enumerate(instance_names):
                marks = []
                for comp, symbol in (("NSGA-II", "*"), ("MOEA/D", "†"), ("HHO", "‡")):
                    if sig_marks.get(inst, {}).get(comp, False):
                        marks.append(symbol)
                if marks:
                    ax.text(
                        x[idx] + (ai - 1.5) * width,
                        means[idx] + (stds[idx] if stds[idx] > 0 else 0.01),
                        "".join(marks),
                        ha="center",
                        va="bottom",
                        fontsize=9,
                    )

    ax.set_xticks(x)
    ax.set_xticklabels(instance_names, rotation=45, ha="right")
    ax.set_ylabel("Contribution Rate")
    ax.set_title("Contribution Rate by Region (EA)")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.legend(title="Region", frameon=False, ncol=3, loc="upper right")
    fig.tight_layout()
    fig.savefig(out_path, dpi=600)
    plt.close(fig)


def _moving_average(values: List[float], window: int) -> List[float]:
    if window <= 1:
        return values
    out = []
    for i in range(len(values)):
        s = 0.0
        c = 0
        for j in range(max(0, i - window + 1), i + 1):
            s += values[j]
            c += 1
        out.append(s / c if c else 0.0)
    return out


def _plot_mode_activation(
    steps: List[int],
    mode_series: Dict[str, List[float]],
    out_path: Path,
    title: str,
) -> None:
    _setup_plot_style()
    fig, ax = plt.subplots(figsize=(6.8, 4.0))
    styles = {
        "independent": ("-", "black"),
        "cooperation": ("--", "dimgray"),
        "competition": ("-.", "gray"),
    }
    for mode, series in mode_series.items():
        ls, color = styles.get(mode, ("-", "black"))
        ax.plot(steps, series, linestyle=ls, color=color, label=mode.capitalize())
    ax.set_xlabel("Iteration")
    ax.set_ylabel("Activation Frequency")
    ax.set_ylim(0.0, 1.0)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.set_title(title)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(out_path, dpi=600)
    plt.close(fig)


def _plot_operator_boxplot(
    data: Dict[str, List[float]],
    out_path: Path,
    title: str,
) -> None:
    _setup_plot_style()
    labels = list(data.keys())
    values = [data[k] for k in labels]
    fig, ax = plt.subplots(figsize=(5.8, 4.0))
    bp = ax.boxplot(
        values,
        labels=labels,
        whis=1.5,
        patch_artist=True,
        showfliers=True,
    )
    for box in bp["boxes"]:
        box.set(facecolor="white", edgecolor="black", linewidth=1.2)
    for med in bp["medians"]:
        med.set(color="black", linewidth=2.0)
    for whisker in bp["whiskers"]:
        whisker.set(color="black", linewidth=1.2)
    for cap in bp["caps"]:
        cap.set(color="black", linewidth=1.2)
    ax.set_ylabel("Success Rate")
    ax.set_title(title)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    fig.tight_layout()
    fig.savefig(out_path, dpi=600)
    plt.close(fig)


def _machines_matrix(rng: random.Random, F: int, S: int) -> List[List[int]]:
    return [[rng.randint(2, 4) for _ in range(S)] for _ in range(F)]


def build_instance(spec: Dict[str, int], seed: int) -> ProblemInstance:
    rng = random.Random(seed)
    J, F, S = spec["J"], spec["F"], spec["S"]
    machines = _machines_matrix(rng, F, S)
    processing_time = [[rng.randint(1, 20) for _ in range(S)] for _ in range(J)]
    due_low = max(10, int(7.5 * S))
    due_high = max(due_low + 1, int(50 * S))
    due_date = [rng.randint(due_low, due_high) for _ in range(J)]
    weight = [rng.randint(1, 5) for _ in range(J)]
    return ProblemInstance(
        num_factories=F,
        num_jobs=J,
        num_stages=S,
        machines=machines,
        processing_time=processing_time,
        due_date=due_date,
        weight=weight,
    )


def _load_archive(path: Path) -> list:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [tuple(x) for x in data.get("objectives", [])]


def _load_summary(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_metrics(metrics: dict) -> dict:
    if not metrics:
        return metrics
    eps_floor = 1e-3
    hv_vals = [m["hv"] for m in metrics.values()]
    igd_vals = [m["igd"] for m in metrics.values()]
    sp_vals = [m["spacing"] for m in metrics.values()]

    max_hv = max(hv_vals) or 1.0
    min_igd, max_igd = min(igd_vals), max(igd_vals)
    min_sp, max_sp = min(sp_vals), max(sp_vals)

    for m in metrics.values():
        hv_norm = (m["hv"] / max_hv) if max_hv > 0 else eps_floor
        igd_norm = (
            (m["igd"] - min_igd) / (max_igd - min_igd + 1e-9)
            if max_igd > min_igd
            else 0.5
        )
        sp_norm = (
            (m["spacing"] - min_sp) / (max_sp - min_sp + 1e-9)
            if max_sp > min_sp
            else 0.5
        )
        m["hv"] = max(hv_norm, eps_floor)
        m["igd"] = max(igd_norm, eps_floor)
        m["spacing"] = max(sp_norm, eps_floor)
    return metrics


def compute_metrics(run_dirs: dict, runtimes: dict) -> dict:
    all_objs = []
    for d in run_dirs.values():
        all_objs.extend(_load_archive(d / "archive.json"))
    ref = nondominated(all_objs)

    results = {}
    for name, d in run_dirs.items():
        objs = _load_archive(d / "archive.json")
        if not objs:
            continue
        f1 = [o[0] for o in objs]
        f2 = [o[1] for o in objs]
        f3 = [o[2] for o in objs]
        hv = hv_approx(objs, (1e4, 1e4, 1e4))
        results[name] = {
            "f1_best": min(f1),
            "f1_mean": sum(f1) / len(f1),
            "f2_best": min(f2),
            "f2_mean": sum(f2) / len(f2),
            "f3_best": min(f3),
            "f3_mean": sum(f3) / len(f3),
            "hv": hv,
            "igd": igd(ref, objs),
            "spacing": spacing(objs),
            "runtime": runtimes.get(name, 0.0),
        }
        if name == "HQ":
            summary = _load_summary(d / "summary.json")
            results[name]["cr"] = summary.get("contribution_rate", {})
            results[name]["mcr"] = summary.get("marginal_contribution_rate", {})
            results[name]["mc_raw"] = summary.get("marginal_contribution", {})
            results[name]["assist_mcr"] = summary.get("assist_marginal_contribution_rate", {})
            results[name]["assist_mc_raw"] = summary.get("assist_marginal_contribution", {})
            results[name]["total_mcr"] = summary.get("total_marginal_contribution_rate", {})
            results[name]["total_mc_raw"] = summary.get("total_marginal_contribution", {})
            results[name]["blended_mcr"] = summary.get("blended_marginal_contribution_rate", {})
            results[name]["blended_mc_raw"] = summary.get("blended_marginal_contribution", {})
            results[name]["mode_counts"] = summary.get("mode_counts", {})
            results[name]["op_counts"] = summary.get("operator_counts", {})
            results[name]["op_success"] = summary.get("operator_success", {})
    return normalize_metrics(results)


def export_text_table(metrics: dict, path: Path) -> None:
    header = [
        "算法",
        "完工时间-最优",
        "完工时间-均值",
        "总拖期-最优",
        "总拖期-均值",
        "负载不均衡-最优",
        "负载不均衡-均值",
        "HV(归一化)",
        "IGD(归一化)",
        "Spacing(归一化)",
        "运行时间(s)",
    ]
    row_fmt = "{:<10} {:>12} {:>12} {:>12} {:>12} {:>14} {:>14} {:>12} {:>12} {:>14} {:>12}"
    lines = []
    lines.append("算法对比（最优/均值/运行时间）")
    lines.append(row_fmt.format(*header))
    for name, m in metrics.items():
        lines.append(
            row_fmt.format(
                name,
                f"{m['f1_best']:.2f}",
                f"{m['f1_mean']:.2f}",
                f"{m['f2_best']:.2f}",
                f"{m['f2_mean']:.2f}",
                f"{m['f3_best']:.2f}",
                f"{m['f3_mean']:.2f}",
                f"{m['hv']:.4f}",
                f"{m['igd']:.4f}",
                f"{m['spacing']:.4f}",
                f"{m['runtime']:.2f}",
            )
        )
    if "HQ" in metrics:
        hq = metrics["HQ"]
        lines.append("")
        lines.append("HQ 机制指标：")
        cr = hq.get("cr", {}) or {}
        mcr = hq.get("mcr", {}) or {}
        mode = hq.get("mode_counts", {}) or {}
        op_counts = hq.get("op_counts", {}) or {}
        op_success = hq.get("op_success", {}) or {}
        mode_total = sum(mode.values()) or 1
        op_rates = {k: (op_success.get(k, 0) / v) for k, v in op_counts.items() if v > 0}
        lines.append(f"- 贡献率(CR): {cr}")
        lines.append(f"- 边际贡献率(MCR): {mcr}")
        lines.append(f"- 协同边际贡献率(Assist-MCR): {hq.get('assist_mcr', {}) or {}}")
        lines.append(f"- 总边际贡献率(Total-MCR): {hq.get('total_mcr', {}) or {}}")
        lines.append(f"- 融合边际贡献率(Blended-MCR): {hq.get('blended_mcr', {}) or {}}")
        lines.append(f"- 模式占比: { {k: v / mode_total for k, v in mode.items()} }")
        lines.append(f"- 算子成功率: {op_rates}")
        lines.append("统计检验：单实例不适用（详见批量汇总）")
    path.write_text("\n".join(lines), encoding="utf-8")


def export_excel(metrics: dict, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "对比表"
    ws.append(
        [
            "算法",
            "完工时间-最优",
            "完工时间-均值",
            "总拖期-最优",
            "总拖期-均值",
            "负载不均衡-最优",
            "负载不均衡-均值",
            "运行时间(s)",
            "HV(归一化)",
            "IGD(归一化)",
            "Spacing(归一化)",
        ]
    )
    for name, m in metrics.items():
        ws.append(
            [
                name,
                m["f1_best"],
                m["f1_mean"],
                m["f2_best"],
                m["f2_mean"],
                m["f3_best"],
                m["f3_mean"],
                m["runtime"],
                m["hv"],
                m["igd"],
                m["spacing"],
            ]
        )
    if "HQ" in metrics:
        ws2 = wb.create_sheet("HQ机制")
        m = metrics["HQ"]
        ws2.append(["指标", "数值"])
        ws2.append(["贡献率(CR)", json.dumps(m.get("cr", {}), ensure_ascii=False)])
        ws2.append(["边际贡献率(MCR)", json.dumps(m.get("mcr", {}), ensure_ascii=False)])
        ws2.append(["边际贡献原值(MC)", json.dumps(m.get("mc_raw", {}), ensure_ascii=False)])
        ws2.append(["协同边际贡献率(Assist-MCR)", json.dumps(m.get("assist_mcr", {}), ensure_ascii=False)])
        ws2.append(["协同边际贡献原值(Assist-MC)", json.dumps(m.get("assist_mc_raw", {}), ensure_ascii=False)])
        ws2.append(["总边际贡献率(Total-MCR)", json.dumps(m.get("total_mcr", {}), ensure_ascii=False)])
        ws2.append(["总边际贡献原值(Total-MC)", json.dumps(m.get("total_mc_raw", {}), ensure_ascii=False)])
        ws2.append(["融合边际贡献率(Blended-MCR)", json.dumps(m.get("blended_mcr", {}), ensure_ascii=False)])
        ws2.append(["融合边际贡献原值(Blended-MC)", json.dumps(m.get("blended_mc_raw", {}), ensure_ascii=False)])
        mode_counts = m.get("mode_counts", {}) or {}
        op_counts = m.get("op_counts", {}) or {}
        op_success = m.get("op_success", {}) or {}
        mode_total = sum(mode_counts.values()) or 1
        mode_rates = {k: v / mode_total for k, v in mode_counts.items()}
        op_rates = {k: (op_success.get(k, 0) / v) for k, v in op_counts.items() if v > 0}
        ws2.append(["模式频率", json.dumps(mode_counts, ensure_ascii=False)])
        ws2.append(["模式占比", json.dumps(mode_rates, ensure_ascii=False)])
        ws2.append(["算子次数", json.dumps(op_counts, ensure_ascii=False)])
        ws2.append(["算子成功", json.dumps(op_success, ensure_ascii=False)])
        ws2.append(["算子成功率", json.dumps(op_rates, ensure_ascii=False)])
        ws3 = wb.create_sheet("统计检验(NA)")
        ws3.append(["说明"])
        ws3.append(["单实例不做显著性检验；请查看批量汇总(batch_metrics.xlsx)"])
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_new{path.suffix}")
        wb.save(fallback)
        return fallback


def _mean_std(values: List[float]) -> Tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return float(values[0]), 0.0
    return float(statistics.mean(values)), float(statistics.stdev(values))


def _aggregate_run_metrics(run_metrics_list: List[dict]) -> dict:
    """Aggregate metrics across repeated runs of the same case."""
    if not run_metrics_list:
        return {}
    algos = ["HQ", "NSGA-II", "MOEA/D", "HHO", "IG"]
    scalar_keys = [
        "f1_best",
        "f1_mean",
        "f2_best",
        "f2_mean",
        "f3_best",
        "f3_mean",
        "hv",
        "igd",
        "spacing",
        "runtime",
    ]
    aggregated: Dict[str, dict] = {}
    for algo in algos:
        vals_by_key: Dict[str, List[float]] = {k: [] for k in scalar_keys}
        cr_vals: Dict[str, List[float]] = {}
        mcr_vecs: List[Dict[str, float]] = []
        assist_mcr_vecs: List[Dict[str, float]] = []
        total_mcr_vecs: List[Dict[str, float]] = []
        blended_mcr_vecs: List[Dict[str, float]] = []
        mode_vals: Dict[str, List[int]] = {}
        op_cnt_vals: Dict[str, List[int]] = {}
        op_suc_vals: Dict[str, List[int]] = {}
        for rm in run_metrics_list:
            m = rm.get(algo, {})
            if not m:
                continue
            for k in scalar_keys:
                if k in m:
                    vals_by_key[k].append(float(m[k]))
            if algo == "HQ":
                for k, v in (m.get("cr", {}) or {}).items():
                    cr_vals.setdefault(k, []).append(float(v))
                mcr = {k: float(v) for k, v in (m.get("mcr", {}) or {}).items()}
                if sum(mcr.values()) > 1e-12:
                    mcr_vecs.append(mcr)
                assist_mcr = {k: float(v) for k, v in (m.get("assist_mcr", {}) or {}).items()}
                if sum(assist_mcr.values()) > 1e-12:
                    assist_mcr_vecs.append(assist_mcr)
                total_mcr = {k: float(v) for k, v in (m.get("total_mcr", {}) or {}).items()}
                if sum(total_mcr.values()) > 1e-12:
                    total_mcr_vecs.append(total_mcr)
                blended_mcr = {k: float(v) for k, v in (m.get("blended_mcr", {}) or {}).items()}
                if sum(blended_mcr.values()) > 1e-12:
                    blended_mcr_vecs.append(blended_mcr)
                for k, v in (m.get("mode_counts", {}) or {}).items():
                    mode_vals.setdefault(k, []).append(int(v))
                for k, v in (m.get("op_counts", {}) or {}).items():
                    op_cnt_vals.setdefault(k, []).append(int(v))
                for k, v in (m.get("op_success", {}) or {}).items():
                    op_suc_vals.setdefault(k, []).append(int(v))

        if any(vals_by_key[k] for k in scalar_keys):
            aggregated[algo] = {}
            for k in scalar_keys:
                if vals_by_key[k]:
                    aggregated[algo][k] = float(statistics.mean(vals_by_key[k]))
            if algo == "HQ":
                aggregated[algo]["cr"] = {k: float(statistics.mean(v)) for k, v in cr_vals.items()}
                if mcr_vecs:
                    keys = ("nsga", "moead", "hho", "ig")
                    mcr_mean = {k: float(statistics.mean(vec.get(k, 0.0) for vec in mcr_vecs)) for k in keys}
                    mcr_sum = sum(mcr_mean.values())
                    if mcr_sum > 1e-12:
                        mcr_mean = {k: v / mcr_sum for k, v in mcr_mean.items()}
                    aggregated[algo]["mcr"] = mcr_mean
                else:
                    aggregated[algo]["mcr"] = {"nsga": 0.0, "moead": 0.0, "hho": 0.0, "ig": 0.0}
                if assist_mcr_vecs:
                    keys = ("nsga", "moead", "hho", "ig")
                    assist_mean = {
                        k: float(statistics.mean(vec.get(k, 0.0) for vec in assist_mcr_vecs))
                        for k in keys
                    }
                    assist_sum = sum(assist_mean.values())
                    if assist_sum > 1e-12:
                        assist_mean = {k: v / assist_sum for k, v in assist_mean.items()}
                    aggregated[algo]["assist_mcr"] = assist_mean
                else:
                    aggregated[algo]["assist_mcr"] = {"nsga": 0.0, "moead": 0.0, "hho": 0.0, "ig": 0.0}
                if total_mcr_vecs:
                    keys = ("nsga", "moead", "hho", "ig")
                    total_mean = {
                        k: float(statistics.mean(vec.get(k, 0.0) for vec in total_mcr_vecs))
                        for k in keys
                    }
                    total_sum = sum(total_mean.values())
                    if total_sum > 1e-12:
                        total_mean = {k: v / total_sum for k, v in total_mean.items()}
                    aggregated[algo]["total_mcr"] = total_mean
                else:
                    aggregated[algo]["total_mcr"] = {"nsga": 0.0, "moead": 0.0, "hho": 0.0, "ig": 0.0}
                if blended_mcr_vecs:
                    keys = ("nsga", "moead", "hho", "ig")
                    blended_mean = {
                        k: float(statistics.mean(vec.get(k, 0.0) for vec in blended_mcr_vecs))
                        for k in keys
                    }
                    blended_sum = sum(blended_mean.values())
                    if blended_sum > 1e-12:
                        blended_mean = {k: v / blended_sum for k, v in blended_mean.items()}
                    aggregated[algo]["blended_mcr"] = blended_mean
                else:
                    aggregated[algo]["blended_mcr"] = {"nsga": 0.0, "moead": 0.0, "hho": 0.0, "ig": 0.0}
                aggregated[algo]["mode_counts"] = {k: int(round(statistics.mean(v))) for k, v in mode_vals.items()}
                aggregated[algo]["op_counts"] = {k: int(round(statistics.mean(v))) for k, v in op_cnt_vals.items()}
                aggregated[algo]["op_success"] = {k: int(round(statistics.mean(v))) for k, v in op_suc_vals.items()}
    return aggregated


def _rank_values(values: List[float], higher_better: bool) -> List[int]:
    # Return ranks (1=best) with average rank for ties
    indexed = list(enumerate(values))
    indexed.sort(key=lambda x: x[1], reverse=higher_better)
    ranks = [0] * len(values)
    i = 0
    while i < len(indexed):
        j = i
        while j < len(indexed) and indexed[j][1] == indexed[i][1]:
            j += 1
        avg_rank = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[indexed[k][0]] = avg_rank
        i = j
    return ranks


def aggregate_case_metrics(case_metrics: List[dict]) -> dict:
    algos = ["HQ", "NSGA-II", "MOEA/D", "HHO", "IG"]
    keys = [
        "f1_best",
        "f1_mean",
        "f2_best",
        "f2_mean",
        "f3_best",
        "f3_mean",
        "hv",
        "igd",
        "spacing",
        "runtime",
    ]
    agg: Dict[str, Dict[str, Tuple[float, float]]] = {a: {} for a in algos}
    for a in algos:
        for k in keys:
            vals = [cm.get(a, {}).get(k) for cm in case_metrics if a in cm and k in cm[a]]
            vals = [v for v in vals if v is not None]
            agg[a][k] = _mean_std([float(v) for v in vals])
    return agg


def aggregate_hq_mechanism(case_metrics: List[dict]) -> dict:
    mode_rates: Dict[str, List[float]] = {}
    op_rates: Dict[str, List[float]] = {}
    cr_rates: Dict[str, List[float]] = {}
    mcr_rates: Dict[str, List[float]] = {}
    assist_mcr_rates: Dict[str, List[float]] = {}
    total_mcr_rates: Dict[str, List[float]] = {}
    blended_mcr_rates: Dict[str, List[float]] = {}

    for cm in case_metrics:
        hq = cm.get("HQ", {})
        mode_counts = hq.get("mode_counts", {}) or {}
        op_counts = hq.get("op_counts", {}) or {}
        op_success = hq.get("op_success", {}) or {}
        cr = hq.get("cr", {}) or {}
        mcr = hq.get("mcr", {}) or {}
        assist_mcr = hq.get("assist_mcr", {}) or {}
        total_mcr = hq.get("total_mcr", {}) or {}
        blended_mcr = hq.get("blended_mcr", {}) or {}

        mode_total = sum(mode_counts.values())
        if mode_total > 0:
            for k, v in mode_counts.items():
                mode_rates.setdefault(k, []).append(v / mode_total)
        for k, v in op_counts.items():
            if v > 0:
                op_rates.setdefault(k, []).append((op_success or {}).get(k, 0) / v)
        for k, v in cr.items():
            cr_rates.setdefault(k, []).append(float(v))
        for k, v in mcr.items():
            mcr_rates.setdefault(k, []).append(float(v))
        for k, v in assist_mcr.items():
            assist_mcr_rates.setdefault(k, []).append(float(v))
        for k, v in total_mcr.items():
            total_mcr_rates.setdefault(k, []).append(float(v))
        for k, v in blended_mcr.items():
            blended_mcr_rates.setdefault(k, []).append(float(v))
        # 确保 nsga/moead/hho/ig 均存在，缺失则记为 0
        for algo in ("nsga", "moead", "hho", "ig"):
            if algo not in cr:
                cr_rates.setdefault(algo, []).append(0.0)
            if algo not in mcr:
                mcr_rates.setdefault(algo, []).append(0.0)
            if algo not in assist_mcr:
                assist_mcr_rates.setdefault(algo, []).append(0.0)
            if algo not in total_mcr:
                total_mcr_rates.setdefault(algo, []).append(0.0)
            if algo not in blended_mcr:
                blended_mcr_rates.setdefault(algo, []).append(0.0)

    return {
        "mode_rates": {k: _mean_std(v) for k, v in mode_rates.items()},
        "op_success_rates": {k: _mean_std(v) for k, v in op_rates.items()},
        "contribution_rates": {k: _mean_std(v) for k, v in cr_rates.items()},
        "marginal_contribution_rates": {k: _mean_std(v) for k, v in mcr_rates.items()},
        "assist_marginal_contribution_rates": {k: _mean_std(v) for k, v in assist_mcr_rates.items()},
        "total_marginal_contribution_rates": {k: _mean_std(v) for k, v in total_mcr_rates.items()},
        "blended_marginal_contribution_rates": {k: _mean_std(v) for k, v in blended_mcr_rates.items()},
    }


def run_stat_tests(case_metrics: List[dict]) -> dict:
    algos = ["HQ", "NSGA-II", "MOEA/D", "HHO", "IG"]
    metrics = {
        "HV": ("hv", True),
        "IGD": ("igd", False),
        "Spacing": ("spacing", False),
    }
    tests = {"wilcoxon": {}, "friedman": {}}
    if wilcoxon is None or friedmanchisquare is None:
        tests["error"] = "scipy not available"
        return tests
    if len(case_metrics) < 2:
        tests["error"] = "insufficient samples (<2 cases)"
        return tests

    for label, (key, higher_better) in metrics.items():
        series = {a: [cm.get(a, {}).get(key) for cm in case_metrics] for a in algos}
        if any(any(v is None for v in vals) for vals in series.values()):
            continue

        # Wilcoxon HQ vs others
        w_res = {}
        for a in algos:
            if a == "HQ":
                continue
            x = [float(v) for v in series["HQ"]]
            y = [float(v) for v in series[a]]
            try:
                stat, p = wilcoxon(x, y, alternative="two-sided", zero_method="wilcox")
            except Exception:
                stat, p = 0.0, 1.0
            w_res[a] = {"stat": float(stat), "p": float(p)}
        tests["wilcoxon"][label] = w_res

        # Friedman + average rank
        try:
            stat, p = friedmanchisquare(*[series[a] for a in algos])
        except Exception:
            stat, p = 0.0, 1.0
        rank_lists = []
        for i in range(len(case_metrics)):
            vals = [float(series[a][i]) for a in algos]
            rank_lists.append(_rank_values(vals, higher_better=higher_better))
        avg_ranks = [statistics.mean(r[i] for r in rank_lists) for i in range(len(algos))]
        tests["friedman"][label] = {
            "stat": float(stat),
            "p": float(p),
            "avg_rank": {algos[i]: float(avg_ranks[i]) for i in range(len(algos))},
        }

    return tests


def export_batch_summary(
    case_metrics: List[dict],
    out_dir: Path,
) -> None:
    agg = aggregate_case_metrics(case_metrics)
    hq_mech = aggregate_hq_mechanism(case_metrics)
    tests = run_stat_tests(case_metrics)

    # text summary
    lines = []
    lines.append("批量统计（均值±标准差）")
    for algo, stats in agg.items():
        lines.append(f"\n[{algo}]")
        for k, (m, s) in stats.items():
            lines.append(f"- {k}: {m:.4f} ± {s:.4f}")
    lines.append("\nHQ 机制（均值±标准差）")
    lines.append(f"- 贡献率: {hq_mech.get('contribution_rates', {})}")
    lines.append(f"- 边际贡献率: {hq_mech.get('marginal_contribution_rates', {})}")
    lines.append(f"- 协同边际贡献率: {hq_mech.get('assist_marginal_contribution_rates', {})}")
    lines.append(f"- 总边际贡献率: {hq_mech.get('total_marginal_contribution_rates', {})}")
    lines.append(f"- 融合边际贡献率: {hq_mech.get('blended_marginal_contribution_rates', {})}")
    lines.append(f"- 模式占比: {hq_mech.get('mode_rates', {})}")
    lines.append(f"- 算子成功率: {hq_mech.get('op_success_rates', {})}")

    lines.append("\n统计检验")
    if "error" in tests:
        lines.append(f"- 错误: {tests['error']}")
    else:
        lines.append("- Wilcoxon (HQ vs others):")
        for metric, res in tests["wilcoxon"].items():
            lines.append(f"  {metric}: {res}")
        lines.append("- Friedman:")
        for metric, res in tests["friedman"].items():
            lines.append(f"  {metric}: stat={res['stat']:.4f}, p={res['p']:.4f}, rank={res['avg_rank']}")

    (out_dir / "batch_metrics.txt").write_text("\n".join(lines), encoding="utf-8")
    (out_dir / "batch_tests.json").write_text(
        json.dumps(tests, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # batch mechanism plots
    cr = hq_mech.get("contribution_rates", {})
    if cr:
        algo_order = [("nsga", "NSGA-II"), ("moead", "MOEA/D"), ("hho", "HHO"), ("ig", "IG")]
        vals = [cr.get(k, (0.0, 0.0))[0] for k, _ in algo_order]
        errs = [cr.get(k, (0.0, 0.0))[1] for k, _ in algo_order]
        labels = [label for _, label in algo_order]
        _plot_bar(
            title="HQ贡献率(CR) - 均值±标准差",
            labels=labels,
            values=vals,
            yerr=errs,
            out_path=out_dir / "batch_hq_mech_cr.pdf",
            ylabel="CR",
        )
    mode_rates = hq_mech.get("mode_rates", {})
    if mode_rates:
        mode_order = [("independent", "Independent"), ("cooperation", "Cooperation"), ("competition", "Competition")]
        vals = [mode_rates.get(k, (0.0, 0.0))[0] for k, _ in mode_order]
        errs = [mode_rates.get(k, (0.0, 0.0))[1] for k, _ in mode_order]
        labels = [label for _, label in mode_order]
        _plot_bar(
            title="HQ模式占比 - 均值±标准差",
            labels=labels,
            values=vals,
            yerr=errs,
            out_path=out_dir / "batch_hq_mech_mode.pdf",
            ylabel="Ratio",
        )
    op_rates = hq_mech.get("op_success_rates", {})
    if op_rates:
        op_order = [
            ("c1_elite_migration", "C1"),
            ("c2_rhythm_coop", "C2"),
            ("r1_struct_suppress", "R1"),
            ("r2_territorial_invade", "R2"),
        ]
        vals = [op_rates.get(k, (0.0, 0.0))[0] for k, _ in op_order]
        errs = [op_rates.get(k, (0.0, 0.0))[1] for k, _ in op_order]
        labels = [label for _, label in op_order]
        _plot_bar(
            title="HQ算子成功率 - 均值±标准差",
            labels=labels,
            values=vals,
            yerr=errs,
            out_path=out_dir / "batch_hq_mech_op_success.pdf",
            ylabel="Success Rate",
        )
    blended_mcr = hq_mech.get("blended_marginal_contribution_rates", {})
    if blended_mcr:
        algo_order = [("nsga", "NSGA-II"), ("moead", "MOEA/D"), ("hho", "HHO"), ("ig", "IG")]
        vals = [blended_mcr.get(k, (0.0, 0.0))[0] for k, _ in algo_order]
        errs = [blended_mcr.get(k, (0.0, 0.0))[1] for k, _ in algo_order]
        labels = [label for _, label in algo_order]
        _plot_bar(
            title="HQ融合边际贡献率(Blended-MCR) - 均值±标准差",
            labels=labels,
            values=vals,
            yerr=errs,
            out_path=out_dir / "batch_hq_mech_blended_mcr.pdf",
            ylabel="Blended-MCR",
        )

    # excel summary
    wb = Workbook()
    ws = wb.active
    ws.title = "批量统计"
    ws.append(["算法", "指标", "均值", "标准差"])
    for algo, stats in agg.items():
        for k, (m, s) in stats.items():
            ws.append([algo, k, m, s])

    ws2 = wb.create_sheet("HQ机制汇总")
    ws2.append(["指标", "子项", "均值", "标准差"])
    for k, v in hq_mech.get("contribution_rates", {}).items():
        ws2.append(["贡献率", k, v[0], v[1]])
    for k, v in hq_mech.get("marginal_contribution_rates", {}).items():
        ws2.append(["边际贡献率", k, v[0], v[1]])
    for k, v in hq_mech.get("assist_marginal_contribution_rates", {}).items():
        ws2.append(["协同边际贡献率", k, v[0], v[1]])
    for k, v in hq_mech.get("total_marginal_contribution_rates", {}).items():
        ws2.append(["总边际贡献率", k, v[0], v[1]])
    for k, v in hq_mech.get("blended_marginal_contribution_rates", {}).items():
        ws2.append(["融合边际贡献率", k, v[0], v[1]])
    for k, v in hq_mech.get("mode_rates", {}).items():
        ws2.append(["模式占比", k, v[0], v[1]])
    for k, v in hq_mech.get("op_success_rates", {}).items():
        ws2.append(["算子成功率", k, v[0], v[1]])

    ws3 = wb.create_sheet("统计检验")
    ws3.append(["检验", "指标", "对象", "stat", "p值", "平均排名"])
    if "error" in tests:
        ws3.append(["错误", "-", tests["error"], "", "", ""])
    else:
        for metric, res in tests["wilcoxon"].items():
            for algo, info in res.items():
                ws3.append(["Wilcoxon", metric, f"HQ vs {algo}", info["stat"], info["p"], ""])
        for metric, res in tests["friedman"].items():
            ws3.append(["Friedman", metric, "all", res["stat"], res["p"], json.dumps(res["avg_rank"])])

    out_path = out_dir / "batch_metrics.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        wb.save(out_path.with_name(f"{out_path.stem}_new{out_path.suffix}"))


def _case_specs() -> List[Dict[str, int]]:
    return [
        {"J": 30, "F": 2, "S": 3},
        {"J": 40, "F": 2, "S": 3},
        {"J": 50, "F": 2, "S": 4},
        {"J": 60, "F": 2, "S": 4},
        {"J": 70, "F": 3, "S": 3},
        {"J": 80, "F": 3, "S": 4},
        {"J": 90, "F": 3, "S": 4},
        {"J": 100, "F": 4, "S": 4},
        {"J": 110, "F": 4, "S": 4},
        {"J": 120, "F": 4, "S": 5},
        {"J": 130, "F": 3, "S": 5},
        {"J": 140, "F": 4, "S": 5},
        {"J": 150, "F": 4, "S": 4},
        {"J": 160, "F": 4, "S": 5},
        {"J": 170, "F": 3, "S": 5},
        {"J": 180, "F": 4, "S": 5},
        {"J": 190, "F": 4, "S": 4},
        {"J": 200, "F": 4, "S": 5},
        {"J": 210, "F": 4, "S": 5},
        {"J": 220, "F": 4, "S": 5},
    ]


def run_one_case(
    inst: ProblemInstance,
    cfg_hq: GlobalConfig,
    cfg_single: GlobalConfig,
    case_dir: Path,
    runs_per_instance: int = 2,
) -> tuple[dict, list, dict, dict, dict]:
    case_dir.mkdir(parents=True, exist_ok=True)
    inst.save_json(case_dir / "instance.json")
    cr_runs: List[Dict[str, Dict[str, float]]] = []
    mode_runs: List[Dict[str, List[int]]] = []
    op_runs: List[Dict[str, float]] = []
    run_metrics_list: List[dict] = []

    for r in range(runs_per_instance):
        run_dir = case_dir / f"run_{r:02d}"
        run_dir.mkdir(parents=True, exist_ok=True)

        # HQ
        out_hq = run_dir / "hq"
        t0 = time.perf_counter()
        arch_hq = run(
            inst,
            cfg_hq,
            RunnerConfig(output_dir=str(out_hq), log_every=1, verbose=True),
        )
        t1 = time.perf_counter()
        runtime_hq = t1 - t0
        if ENABLE_RUN_PLOTS:
            plot_progress(out_hq / "progress.jsonl", out_hq)
            plot_pareto([e.objectives for e in arch_hq.entries], out_hq / "pareto.png")
            if arch_hq.entries:
                best = min(arch_hq.entries, key=lambda e: sum(e.objectives))
                plot_gantt(decode(inst, best.encoding).schedule, out_hq / "gantt.png")

        # Singles
        runtimes = {"HQ": runtime_hq}
        for algo in ("nsga", "moead", "hho", "ig"):
            out_dir = run_dir / algo
            t0 = time.perf_counter()
            arch = run_single(
                inst,
                cfg_single,
                RunnerConfig(output_dir=str(out_dir), log_every=20, verbose=True),
                algo,
            )
            t1 = time.perf_counter()
            runtimes[algo.upper() if algo != "moead" else "MOEA/D"] = t1 - t0
            if ENABLE_RUN_PLOTS:
                plot_progress(out_dir / "progress.jsonl", out_dir)
                plot_pareto([e.objectives for e in arch.entries], out_dir / "pareto.png")
                if arch.entries:
                    best = min(arch.entries, key=lambda e: sum(e.objectives))
                    plot_gantt(decode(inst, best.encoding).schedule, out_dir / "gantt.png")

        if ENABLE_RUN_PLOTS:
            plot_hv_compare(
                {
                    "hq": run_dir / "hq",
                    "nsga": run_dir / "nsga",
                    "moead": run_dir / "moead",
                    "hho": run_dir / "hho",
                    "ig": run_dir / "ig",
                },
                run_dir / "hv_compare.png",
            )

            title = f"{inst.num_jobs}J{inst.num_stages}S{inst.num_factories}F - 帕累托前沿对比"
            plot_pareto_compare(
                {
                    "HQ": run_dir / "hq",
                    "NSGA-II": run_dir / "nsga",
                    "MOEA/D": run_dir / "moead",
                    "HHO": run_dir / "hho",
                    "IG": run_dir / "ig",
                },
                title,
                run_dir / "pareto_compare.png",
            )

        run_dirs = {
            "HQ": run_dir / "hq",
            "NSGA-II": run_dir / "nsga",
            "MOEA/D": run_dir / "moead",
            "HHO": run_dir / "hho",
            "IG": run_dir / "ig",
        }
        metrics = compute_metrics(run_dirs, runtimes)
        export_text_table(metrics, run_dir / "comparison.txt")
        export_excel(metrics, run_dir / "comparison.xlsx")
        plot_hq_mechanism_case(metrics, run_dir, run_dir.name)
        run_metrics_list.append(metrics)

        algo_objs = {
            "HQ": _load_archive(run_dir / "hq" / "archive.json"),
            "NSGA-II": _load_archive(run_dir / "nsga" / "archive.json"),
            "MOEA/D": _load_archive(run_dir / "moead" / "archive.json"),
            "HHO": _load_archive(run_dir / "hho" / "archive.json"),
        }
        cr_runs.append(_contribution_by_region(algo_objs))

        mode_series: Dict[str, List[int]] = {"independent": [], "cooperation": [], "competition": []}
        if (out_hq / "progress.jsonl").exists():
            with (out_hq / "progress.jsonl").open("r", encoding="utf-8") as f:
                for line in f:
                    obj = json.loads(line)
                    mode = obj.get("mode")
                    for k in mode_series.keys():
                        mode_series[k].append(1 if mode == k else 0)
        mode_runs.append(mode_series)

        summary = _load_summary(out_hq / "summary.json")
        op_counts = summary.get("operator_counts", {}) or {}
        op_success = summary.get("operator_success", {}) or {}
        op_rates = {}
        for k, v in op_counts.items():
            if v > 0:
                op_rates[k] = op_success.get(k, 0) / v
        op_runs.append(op_rates)

    instance_name = case_dir.name
    sig_marks: Dict[str, Dict[str, bool]] = {instance_name: {}}
    if wilcoxon is not None:
        comp_list = ["NSGA-II", "MOEA/D", "HHO"]
        for comp in comp_list:
            hq_vals = [
                r.get("HQ", {}).get("extreme", 0)
                + r.get("HQ", {}).get("middle", 0)
                + r.get("HQ", {}).get("knee", 0)
                for r in cr_runs
            ]
            comp_vals = [
                r.get(comp, {}).get("extreme", 0)
                + r.get(comp, {}).get("middle", 0)
                + r.get(comp, {}).get("knee", 0)
                for r in cr_runs
            ]
            if hq_vals and comp_vals:
                try:
                    stat, p = wilcoxon(hq_vals, comp_vals, alternative="two-sided", zero_method="wilcox")
                except Exception:
                    p = 1.0
                sig_marks[instance_name][comp] = p < 0.05 and statistics.mean(hq_vals) > statistics.mean(comp_vals)

    if ENABLE_CASE_SUMMARY_PLOTS:
        _plot_cr_grouped_stacked(
            [instance_name],
            {instance_name: cr_runs},
            case_dir / "hq_cr_region.pdf",
            sig_marks,
        )

    if mode_runs:
        steps = list(range(len(mode_runs[0]["independent"])))
        avg_modes = {}
        for k in ("independent", "cooperation", "competition"):
            series = [sum(run[k][i] for run in mode_runs) / len(mode_runs) for i in range(len(steps))]
            avg_modes[k] = _moving_average(series, 5)
        if ENABLE_CASE_SUMMARY_PLOTS:
            _plot_mode_activation(
                steps,
                avg_modes,
                case_dir / "hq_mode_activation.pdf",
                f"{instance_name} - Mode Activation Frequency",
            )

    if op_runs:
        op_data: Dict[str, List[float]] = {}
        for k in ("c1_elite_migration", "c2_rhythm_coop", "r1_struct_suppress", "r2_territorial_invade"):
            vals = [r.get(k) for r in op_runs if k in r]
            if vals:
                op_data[k] = vals
        if op_data:
            if ENABLE_CASE_SUMMARY_PLOTS:
                _plot_operator_boxplot(
                    op_data,
                    case_dir / "hq_op_success_box.pdf",
                    f"{instance_name} - Operator Success Rate",
                )

    metrics = _aggregate_run_metrics(run_metrics_list)

    print("\n结果输出路径：")
    print(f"- 运行目录：{case_dir}")
    print(f"- HQ机制(CR/模式/算子)：{case_dir / 'hq_cr_region.pdf'}, {case_dir / 'hq_mode_activation.pdf'}, {case_dir / 'hq_op_success_box.pdf'}")
    return metrics, cr_runs, sig_marks, mode_runs, op_runs


def main() -> None:
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    base_out = Path("outputs") / f"compare_batch_{timestamp}"
    base_out.mkdir(parents=True, exist_ok=True)

    max_iters = MEDIUM_MAX_ITERS if QUICK_MEDIUM_TEST else 100
    blend_lambda_env = os.environ.get("ASSIST_BLEND_LAMBDA", "").strip()
    blend_lambda_override = float(blend_lambda_env) if blend_lambda_env else None

    cfg_hq = GlobalConfig(
        population_size=100,
        max_iters=max_iters,
        nsga2=NSGA2Config(population_size=25),
        moead=MOEADConfig(num_subproblems=25),
    )
    cfg_single = GlobalConfig(
        population_size=100,
        max_iters=max_iters,
        nsga2=NSGA2Config(population_size=100),
        moead=MOEADConfig(num_subproblems=100),
    )
    if blend_lambda_override is not None:
        # Keep all other low-level settings unchanged; only sweep blend lambda.
        low_cfg = cfg_hq.low_level
        cfg_hq = GlobalConfig(
            population_size=cfg_hq.population_size,
            max_iters=cfg_hq.max_iters,
            nsga2=cfg_hq.nsga2,
            moead=cfg_hq.moead,
            hho=cfg_hq.hho,
            ig=cfg_hq.ig,
            high_level=cfg_hq.high_level,
            low_level=type(low_cfg)(
                alpha=low_cfg.alpha,
                gamma=low_cfg.gamma,
                epsilon=low_cfg.epsilon,
                mc_weight=low_cfg.mc_weight,
                mc_smooth=low_cfg.mc_smooth,
                target_share=low_cfg.target_share,
                fairness_boost=low_cfg.fairness_boost,
                early_ig_penalty=low_cfg.early_ig_penalty,
                early_phase_ratio=low_cfg.early_phase_ratio,
                assist_boost=low_cfg.assist_boost,
                assist_decay=low_cfg.assist_decay,
                assist_delay_steps=low_cfg.assist_delay_steps,
                assist_reward_scale=low_cfg.assist_reward_scale,
                assist_winner_share=low_cfg.assist_winner_share,
                assist_loser_share=low_cfg.assist_loser_share,
                assist_blend_lambda=blend_lambda_override,
                weak_share_threshold=low_cfg.weak_share_threshold,
                hho_priority_boost=low_cfg.hho_priority_boost,
                hho_competition_protect=low_cfg.hho_competition_protect,
                hho_coop_absorb_rate_scale=low_cfg.hho_coop_absorb_rate_scale,
                hho_compete_loss_rate_scale=low_cfg.hho_compete_loss_rate_scale,
                hho_pair_c2_bias=low_cfg.hho_pair_c2_bias,
                hho_pair_r2_bias=low_cfg.hho_pair_r2_bias,
                hho_stagnation_c2_bonus=low_cfg.hho_stagnation_c2_bonus,
                hho_stagnation_r2_bonus=low_cfg.hho_stagnation_r2_bonus,
                hho_compete_win_rate_scale=low_cfg.hho_compete_win_rate_scale,
            ),
        )
        print(f"\n[参数覆盖] ASSIST_BLEND_LAMBDA={blend_lambda_override:.2f}")

    if QUICK_MEDIUM_TEST:
        specs = [MEDIUM_SPEC]
        runs_per = MEDIUM_RUNS
        print(f"\n[快速中规模] 仅测试 {MEDIUM_SPEC['J']}J{MEDIUM_SPEC['S']}S{MEDIUM_SPEC['F']}F，共 {runs_per} 次运行")
    else:
        specs = _case_specs()
        runs_per = 3

    summary = []
    case_metrics: List[dict] = []
    instance_names: List[str] = []
    cr_runs_map: Dict[str, List[Dict[str, Dict[str, float]]]] = {}
    sig_marks_map: Dict[str, Dict[str, bool]] = {}
    mode_runs_all: List[Dict[str, List[int]]] = []
    op_runs_all: List[Dict[str, float]] = []
    for i, spec in enumerate(specs, start=1):
        if not QUICK_MEDIUM_TEST and CASE_LIMIT > 0 and i > CASE_LIMIT:
            break
        case_name = f"case_{i:02d}_{spec['J']}J{spec['S']}S{spec['F']}F"
        case_dir = base_out / case_name
        seed = 1000 + i

        print("\n==============================")
        print(f"开始测试：{case_name}")
        print(f"规模详情：J={spec['J']} F={spec['F']} S={spec['S']}，每例 {runs_per} 次运行")
        print("==============================")

        inst = build_instance(spec, seed)
        metrics, cr_runs, sig_marks, mode_runs, op_runs = run_one_case(
            inst, cfg_hq, cfg_single, case_dir, runs_per_instance=runs_per
        )
        case_metrics.append(metrics)
        instance_names.append(case_name)
        cr_runs_map[case_name] = cr_runs
        sig_marks_map.update(sig_marks)
        mode_runs_all.extend(mode_runs)
        op_runs_all.extend(op_runs)

        summary.append(
            {
                "case": case_name,
                "seed": seed,
                "J": spec["J"],
                "F": spec["F"],
                "S": spec["S"],
                "output_dir": str(case_dir),
                "instance": str(case_dir / "instance.json"),
            }
        )

    (base_out / "batch_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    export_batch_summary(case_metrics, base_out)

    # Batch CR grouped stacked plot
    if instance_names:
        _plot_cr_grouped_stacked(
            instance_names,
            cr_runs_map,
            base_out / "batch_cr_region.pdf",
            sig_marks_map,
        )

    # Batch mode activation (average over all runs)
    if mode_runs_all:
        steps = list(range(len(mode_runs_all[0]["independent"])))
        avg_modes = {}
        for k in ("independent", "cooperation", "competition"):
            series = [sum(run[k][i] for run in mode_runs_all) / len(mode_runs_all) for i in range(len(steps))]
            avg_modes[k] = _moving_average(series, 5)
        _plot_mode_activation(
            steps,
            avg_modes,
            base_out / "batch_mode_activation.pdf",
            "Mode Activation Frequency (Average)",
        )

    # Batch operator success rate boxplot
    if op_runs_all:
        op_data: Dict[str, List[float]] = {}
        for k in ("c1_elite_migration", "c2_rhythm_coop", "r1_struct_suppress", "r2_territorial_invade"):
            vals = [r.get(k) for r in op_runs_all if k in r]
            if vals:
                op_data[k] = vals
        if op_data:
            _plot_operator_boxplot(
                op_data,
                base_out / "batch_op_success_box.pdf",
                "Operator Success Rate (All Runs)",
            )

    print("\n全部测试完成。")
    print(f"- 批次目录：{base_out}")
    print(f"- 批次汇总：{base_out / 'batch_summary.json'}")
    print(f"- 批量统计：{base_out / 'batch_metrics.xlsx'}")
    print(f"- 批量统计文本：{base_out / 'batch_metrics.txt'}")
    print(f"- 统计检验JSON：{base_out / 'batch_tests.json'}")
    print(
        f"- HQ机制汇总图：{base_out / 'batch_hq_mech_cr.pdf'}, {base_out / 'batch_hq_mech_mode.pdf'}, "
        f"{base_out / 'batch_hq_mech_op_success.pdf'}, {base_out / 'batch_hq_mech_blended_mcr.pdf'}"
    )
    print(f"- CR结构图：{base_out / 'batch_cr_region.pdf'}")
    print(f"- 模式激活曲线：{base_out / 'batch_mode_activation.pdf'}")
    print(f"- 算子成功率箱线图：{base_out / 'batch_op_success_box.pdf'}")


if __name__ == "__main__":
    main()
